"""Analyze triage xlsx — works with files exceljs can't parse."""
import sys
from openpyxl import load_workbook

VALID_CLOVER = {
    "Habitat", "Substrat", "Decoration", "Roche Et Bois/Cork", "Plante Vivant",
    "Lumiere", "Piece Remplacement/ Electronique", "Insecte Vivant", "Nourriture",
    "Supplement", "Serpent", "Lezard", "Rongeur", "Amphibien",
    "Isopod et Araignee et Fourmis",
}
VALID_SITE = {
    "terrarium", "substrate", "decor", "lighting", "equipment",
    "food_live", "food_frozen", "food_packaged", "supplement",
}
VALID_TIPO = {"product", "animal", "ignorar"}


def load_main_sheet(path: str):
    wb = load_workbook(path, data_only=True)
    ws = next(
        s for name, s in ((n, wb[n]) for n in wb.sheetnames)
        if "Sin categor" in name or "588" in name
    )
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h else "" for h in headers]

    def idx(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    def idx_inc(part: str) -> int:
        for i, h in enumerate(headers):
            if part in h:
                return i
        return -1

    cols = {
        "id": idx("cloverItemId"),
        "name": idx("name"),
        "status": idx("queue_status"),
        "conf": idx("confidence"),
        "sug_clover": idx("sugerencia_clover"),
        "sug_site": idx("sugerencia_sitio"),
        "sug_tipo": idx("sugerencia_tipo"),
        "k": idx_inc("ELEGIR categoria Clover"),
        "l": idx_inc("ELEGIR categoria sitio"),
        "m": idx_inc("ELEGIR tipo"),
        "n": idx("notas"),
    }

    rows = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, cols["id"] + 1).value
        if not cid:
            continue
        row = {
            "r": r,
            "id": str(cid),
            "name": str(ws.cell(r, cols["name"] + 1).value or ""),
            "status": str(ws.cell(r, cols["status"] + 1).value or ""),
            "conf": str(ws.cell(r, cols["conf"] + 1).value or ""),
            "clover": str(ws.cell(r, cols["k"] + 1).value or "").strip(),
            "site": str(ws.cell(r, cols["l"] + 1).value or "").strip(),
            "tipo": str(ws.cell(r, cols["m"] + 1).value or "").strip(),
            "notas": str(ws.cell(r, cols["n"] + 1).value or "").strip(),
        }
        rows.append(row)
    return wb.sheetnames, headers, rows


def analyze(path: str, compare_v2: str | None = None):
    sheets, headers, rows = load_main_sheet(path)
    print(f"File: {path}")
    print(f"Sheets: {sheets}")
    print(f"Total rows: {len(rows)}")

    missing_k = [x for x in rows if not x["clover"]]
    missing_m = [x for x in rows if not x["tipo"]]
    missing_l_product = [x for x in rows if x["tipo"] == "product" and not x["site"]]
    should_empty_l = [x for x in rows if x["tipo"] in ("animal", "ignorar") and x["site"]]
    invalid_clover = [x for x in rows if x["clover"] and x["clover"] not in VALID_CLOVER]
    invalid_site = [x for x in rows if x["site"] and x["site"] not in VALID_SITE]
    invalid_tipo = [x for x in rows if x["tipo"] and x["tipo"] not in VALID_TIPO]

    by_tipo: dict[str, int] = {}
    by_clover: dict[str, int] = {}
    for x in rows:
        by_tipo[x["tipo"] or "(vacío)"] = by_tipo.get(x["tipo"] or "(vacío)", 0) + 1
        if x["clover"]:
            by_clover[x["clover"]] = by_clover.get(x["clover"], 0) + 1

    ignorar = [x for x in rows if x["tipo"] == "ignorar"]
    animals = [x for x in rows if x["tipo"] == "animal"]
    pt_habitat = [
        x for x in rows
        if x["name"].upper().startswith("PT") and x["clover"] == "Habitat" and x["tipo"] == "product"
    ]

    print("\n=== RESUMEN ===")
    print(f"K llenas: {len(rows) - len(missing_k)}/{len(rows)}")
    print(f"M llenas: {len(rows) - len(missing_m)}/{len(rows)}")
    print(f"Por tipo: {by_tipo}")
    print(f"ignorar: {len(ignorar)} | animal: {len(animals)} | product: {by_tipo.get('product', 0)}")
    print(f"Product sin L: {len(missing_l_product)}")
    print(f"Animal/ignorar con L llena: {len(should_empty_l)}")
    print(f"Clover inválidos: {len(invalid_clover)}")
    print(f"Sitio inválidos: {len(invalid_site)}")
    print(f"Tipo inválidos: {len(invalid_tipo)}")
    print(f"PT→Habitat sospechosos: {len(pt_habitat)}")

    print("\nTop Clover:")
    for k, v in sorted(by_clover.items(), key=lambda kv: -kv[1]):
        print(f"  {v}\t{k}")

    def show(label, arr, n=12):
        if not arr:
            return
        print(f"\n{label} ({len(arr)}):")
        for x in arr[:n]:
            print(f"  {x['r']}: {x['name'][:52]} | K={x['clover']} L={x['site']} M={x['tipo']}")

    show("Sin K", missing_k)
    show("Sin M", missing_m)
    show("Product sin L", missing_l_product)
    show("Animal/ignorar con L", should_empty_l)
    show("Clover inválido", invalid_clover)
    show("Sitio inválido", invalid_site)
    show("Tipo inválido", invalid_tipo)
    show("ignorar", ignorar, 15)
    show("PT→Habitat (ejemplos)", pt_habitat, 8)

    if compare_v2:
        _, _, v2rows = load_main_sheet(compare_v2)
        v2map = {x["id"]: x for x in v2rows}
        changed_prefilled = []
        newly_filled = []
        for x in rows:
            prev = v2map.get(x["id"])
            if not prev:
                continue
            had_k = bool(prev["clover"])
            has_k = bool(x["clover"])
            if had_k and (prev["clover"] != x["clover"] or prev["site"] != x["site"] or prev["tipo"] != x["tipo"]):
                changed_prefilled.append((prev, x))
            if not had_k and has_k:
                newly_filled.append(x)
        print(f"\n=== vs {compare_v2} ===")
        print(f"Filas que tenían K en v2: {sum(1 for x in v2rows if x['clover'])}")
        print(f"Nuevas completadas por ChatGPT: {len(newly_filled)}")
        print(f"v2 auto-relleno alterado por ChatGPT: {len(changed_prefilled)}")
        show("Auto-relleno modificado (NO debía)", [b for _, b in changed_prefilled], 10)
        show("Recién completadas (ejemplos)", newly_filled, 12)


if __name__ == "__main__":
    path = sys.argv[1]
    cmp = sys.argv[2] if len(sys.argv) > 2 else "exports/chatgpt-v2.xlsx"
    analyze(path, cmp)
