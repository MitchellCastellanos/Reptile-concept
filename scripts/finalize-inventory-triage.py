#!/usr/bin/env python3
"""
Finalize inventory triage:
1. Read chatgpt-v2-completado-final.xlsx
2. Fix misclassified Exo Terra PT#### items (Habitat -> correct category)
3. Write exports/inventaire-final.xlsx + Clover checklist CSVs

Run: npm run inventory:finalize
"""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "exports"
SOURCES = [
    ROOT / "chatgpt-v2-completado-final.xlsx",
    EXPORTS / "chatgpt-v2-completado-final.xlsx",
    EXPORTS / "inventaire-final.xlsx",
]
OUT_XLSX = EXPORTS / "inventaire-final.xlsx"
OUT_MASTER = EXPORTS / "clover-checklist-master.csv"
OUT_DIR = EXPORTS / "clover-checklist"

SITE_BY_CLOVER = {
    "Habitat": "terrarium",
    "Substrat": "substrate",
    "Decoration": "decor",
    "Roche Et Bois/Cork": "decor",
    "Plante Vivant": "decor",
    "Lumiere": "lighting",
    "Piece Remplacement/ Electronique": "equipment",
    "Insecte Vivant": "food_live",
    "Nourriture": "food_packaged",
    "Supplement": "supplement",
}


def find_source() -> Path:
    for p in SOURCES:
        if p.exists():
            return p
    raise FileNotFoundError("No se encontró chatgpt-v2-completado-final.xlsx")


def fix_pt(name: str, clover: str, site: str, tipo: str) -> tuple[str, str, bool]:
    if tipo != "product" or not re.match(r"^PT\d", name, re.I):
        return clover, site, False
    if clover != "Habitat":
        return clover, site, False
    n = name.lower()
    if re.search(
        r"suppl|calcium|electrolyte|vitamin|aquatize|traitement|repti.?boost|electro",
        n,
    ):
        return "Supplement", "supplement", True
    if re.search(
        r"uvb|ampoule|lampe|bulb|solar glo|heat|basking|ceramic|nightlight|daylight|mercury",
        n,
    ):
        return "Lumiere", "lighting", True
    if re.search(
        r"mist|brouillard|ultrason|pump|filter|thermo|hygro|timer|controller|pince|humidif",
        n,
    ):
        return "Piece Remplacement/ Electronique", "equipment", True
    if re.search(
        r"grotte|hide|cachette|water dish|bol d'eau|mousse|substrat|decor|branch|feuille",
        n,
    ):
        return "Decoration", "decor", True
    if re.search(
        r"terrarium|faunarium|paludarium|habitat|screen terrarium|reptibreeze|natural terrarium",
        n,
    ):
        return clover, site, False
    return clover, site, False


def clean_display_name(name: str) -> str:
    cleaned = name.strip()
    cleaned = re.sub(r"^[A-Z0-9]{2,6}\s+(?=[A-Za-z])", "", cleaned)
    cleaned = re.sub(r"^\d{5}\s+", "", cleaned)
    return cleaned.strip()


def load_rows(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = next(w for n, w in ((n, wb[n]) for n in wb.sheetnames) if "Sin categor" in n)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h else "" for h in headers]

    def idx(name: str) -> int:
        return headers.index(name)

    def idx_inc(part: str) -> int:
        for i, h in enumerate(headers):
            if part in h:
                return i
        raise ValueError(f"Columna no encontrada: {part}")

    cols = {
        "id": idx("cloverItemId"),
        "name": idx("name"),
        "price": idx("priceCAD"),
        "stock": idx("stockCount"),
        "status": idx("queue_status"),
        "k": idx_inc("ELEGIR categoria Clover"),
        "l": idx_inc("ELEGIR categoria sitio"),
        "m": idx_inc("ELEGIR tipo"),
        "n": idx("notas"),
    }

    rows = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, cols["id"] + 1).value
        if not cid:
            continue
        rows.append(
            {
                "row": r,
                "ws": ws,
                "cols": cols,
                "id": str(cid),
                "name": str(ws.cell(r, cols["name"] + 1).value or ""),
                "price": ws.cell(r, cols["price"] + 1).value,
                "stock": ws.cell(r, cols["stock"] + 1).value,
                "status": str(ws.cell(r, cols["status"] + 1).value or ""),
                "clover": str(ws.cell(r, cols["k"] + 1).value or "").strip(),
                "site": str(ws.cell(r, cols["l"] + 1).value or "").strip(),
                "tipo": str(ws.cell(r, cols["m"] + 1).value or "").strip(),
                "notas": str(ws.cell(r, cols["n"] + 1).value or "").strip(),
            }
        )
    return wb, ws, headers, rows


def main():
    source = find_source()
    print(f"Fuente: {source}")
    wb, ws, headers, rows = load_rows(source)

    pt_fixed = 0
    for row in rows:
        new_clover, new_site, changed = fix_pt(
            row["name"], row["clover"], row["site"], row["tipo"]
        )
        if changed:
            c = row["cols"]
            r = row["row"]
            note = "v3: PT corregido auto"
            existing = row["notas"]
            row["notas"] = f"{existing}; {note}".strip("; ").strip() if existing else note
            row["clover"] = new_clover
            row["site"] = new_site
            ws.cell(r, c["k"] + 1).value = new_clover
            ws.cell(r, c["l"] + 1).value = new_site
            ws.cell(r, c["n"] + 1).value = row["notas"]
            pt_fixed += 1

    # Sync site category from clover when product and L empty
    for row in rows:
        if row["tipo"] == "product" and row["clover"] and not row["site"]:
            site = SITE_BY_CLOVER.get(row["clover"], "")
            if site:
                row["site"] = site
                ws.cell(row["row"], row["cols"]["l"] + 1).value = site

    EXPORTS.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb.save(OUT_XLSX)
    print(f"Guardado: {OUT_XLSX} (PT corregidos: {pt_fixed})")

    # Checklist CSVs for Clover POS work
    by_clover: dict[str, list] = defaultdict(list)
    checklist_rows = []

    for row in rows:
        if row["tipo"] == "ignorar":
            continue
        clover = row["clover"]
        if not clover:
            continue
        clean = clean_display_name(row["name"])
        entry = {
            "cloverItemId": row["id"],
            "nombre_actual_recibo": row["name"],
            "nombre_sugerido_web": clean,
            "asignar_categoria_clover": clover,
            "categoria_sitio": row["site"] if row["tipo"] == "product" else "",
            "tipo": row["tipo"],
            "queue_status": row["status"],
            "precio_cad": row["price"],
            "accion_clover": f"Asigner categorie: {clover}",
        }
        checklist_rows.append(entry)
        by_clover[clover].append(entry)

    fieldnames = list(checklist_rows[0].keys()) if checklist_rows else []
    with OUT_MASTER.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(sorted(checklist_rows, key=lambda x: (x["asignar_categoria_clover"], x["nombre_actual_recibo"])))

    for clover, items in sorted(by_clover.items()):
        safe = re.sub(r'[<>:"/\\|?*]', "-", clover)[:60]
        path = OUT_DIR / f"{safe}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(sorted(items, key=lambda x: x["nombre_actual_recibo"]))

    ignorar = sum(1 for r in rows if r["tipo"] == "ignorar")
    products = sum(1 for r in rows if r["tipo"] == "product")
    animals = sum(1 for r in rows if r["tipo"] == "animal")
    clover_assign = sum(1 for r in rows if r["tipo"] != "ignorar" and r["clover"])

    print(
        {
            "total": len(rows),
            "ignorar": ignorar,
            "product": products,
            "animal": animals,
            "para_clover": clover_assign,
            "csvs_por_categoria": len(by_clover),
            "master_csv": str(OUT_MASTER),
        }
    )


if __name__ == "__main__":
    main()
